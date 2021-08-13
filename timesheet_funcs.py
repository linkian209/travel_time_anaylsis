'''timesheet_funcs.py

This contains the functions used by the main script
'''
import openpyxl


def print_usage():
    '''
    This function prints out the scripts usage.
    '''
    print('Expected Usage:')
    print('timesheet.py [timesheet excel path] [results excel path]\n')
    print('Required Positional Arguments:')
    print('1: Timesheet Excel Path - Path to Timesheet Excel Doc')
    print('2: Results Excel Path - Path to results Excel Doc')

def make_new_year_page(workbook, year):
    '''
    This function creates a new page for storing results

    Arguments:
        :param workbook: (openpyxl.workbook) The object that we will write into
        :param year: (int) The year
    '''
    # Variables
    bold = openpyxl.styles.Font(b=True)

    # First lets make the new sheet at the end
    workbook.create_sheet(title=str(year))
    sheet = workbook.get_sheet_by_name(str(year))

    # Next lets do the cell merging, styling, and valuing
    sheet.merge_cells('A1:F1')
    cell = sheet['A1']
    cell.fill = openpyxl.styles.PatternFill(patternType="solid", fgColor="dadada")
    cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    cell.font = bold
    cell.value = '{} Totals'.format(year)
    
    sheet.merge_cells('A16:F16')
    cell = sheet['A16']
    cell.fill = openpyxl.styles.PatternFill(patternType="solid", fgColor="dadada")
    cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    cell.font = bold
    cell.value = 'Calculations'

    sheet.merge_cells('A19:F19')
    cell = sheet['A19']
    cell.fill = openpyxl.styles.PatternFill(patternType="solid", fgColor="dadada")
    cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    cell.font = bold
    cell.value = 'Totals Breakdown'

    # Next format column widths
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['C'].width = 13
    sheet.column_dimensions['D'].width = 10
    sheet.column_dimensions['E'].width = 17.5
    sheet.column_dimensions['F'].width = 16.5

    # Next bold and value title row and total row
    titles = ['Month','Hours','Days Worked','','Cumulative Hours','Cumulative Days']
    totals = ['Total','','','','=E14','=F14']
    for col in range(6):
        sheet.cell(row=2, column=col+1).font = bold
        sheet.cell(row=2, column=col+1).value = titles[col]
        sheet.cell(row=15, column=col+1).font = bold
        sheet.cell(row=15, column=col+1).value = totals[col]
    
    # Next put in formulas for the cumulative hours and days
    for row in range(3, 15):
        # If we are the first row in this, we have it easy
        if(row is 3):
            sheet.cell(row=row, column=5).value = '=$B$3'
            sheet.cell(row=row, column=6).value = '=$C$3'
        # Else we need to add the new values to the running totals
        else:
            sheet.cell(row=row, column=5).value = '=E{}+B{}'.format(row-1, row)
            sheet.cell(row=row, column=6).value = '=F{}+C{}'.format(row-1, row)

    # Next, fill in calculations section
    sheet['A17'].font = bold
    sheet['A17'].value = 'Days in Year'
    sheet['B17'].value = 260
    sheet['A18'].font = bold
    sheet['A18'].value = 'Hours in Year'
    sheet['B18'].value = 2080
    sheet['D17'].font = bold
    sheet['D17'].value = '% in Days'
    sheet['E17'].value = '=ABS(F14/B17)'
    sheet['E17'].number_format = '0.00%'
    sheet['D18'].font = bold
    sheet['D18'].value = '% in Hours'
    sheet['E18'].value = '=ABS(E14/B18)'
    sheet['E18'].number_format = '0.00%'

    # Finally do break downs
    sheet['A20'].font = bold
    sheet['A20'].value = 'Days Domestic'
    sheet['A21'].font = bold
    sheet['A21'].value = 'Hours Domestic'
    sheet['E20'].font = bold
    sheet['E20'].value = 'Days Foreign'
    sheet['E21'].font = bold
    sheet['E21'].value = 'Hours Foreign'

    # Sort the sheets
    sheetnames = workbook.sheetnames
    sheetnames.sort()
    workbook._sheets = [workbook.get_sheet_by_name(i) for i in sheetnames]