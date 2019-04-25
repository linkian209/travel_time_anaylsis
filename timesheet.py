'''timesheet.py

This script runs through a CTI Excel timesheet and calculates the number of
days and hours worked out of town.
'''
import datetime
import openpyxl
import os
import sys
from timesheet_funcs import print_usage, make_new_year_page


# Globals
TOTAL_SHEET_NAME = 'Totals'
PO_COLUMN = 1
PO_START_ROW = 4
DAYS_START_COLUMN = 3
DAYS_END_COLUMN = 33
DAY_OF_WEEK_ROW = 2
DAY_NUMBER_ROW = 3
HOLIDAY = 'CTI Holiday'
ALT_HOLIDAY = '=January!$A$64'
ALT_ALT_HOLIDAY = '=January!$A$36'
YEAR_CELL = 'AF1'
cur_year = 0

#
# Main Program
#

# Get commandline arguments, there should be 3
if(len(sys.argv) is not 3):
    print('Not enough arguments')
    print_usage()
    sys.exit(1)

# Check if file exists
try:
    timesheet_file = sys.argv[1]
    results_file = sys.argv[2]
except():
    print('Invalid file entered.')
    print_usage()
    sys.exit(1)

if(not os.path.exists(timesheet_file)):
    print('{} does not exist! Must use real file'.format(timesheet_file))
    print_usage()
    sys.exit(1)

# If the results file doesn't exist, make it
if(not os.path.exists(results_file)):
    print('{} does not exist! Creating...'.format(timesheet_file))
    try:
        results = openpyxl.workbook.Workbook()
        results.get_active_sheet()['A1'].value = 'Delete this sheet'
        results.save(results_file)
    except Exception as e:
        print('Encountered error {} creating results file!'.format(e))
        print_usage()
        sys.exit(1)

# Open up timesheet
timesheet = openpyxl.load_workbook(timesheet_file, read_only=True)

# Log where we are
print('Timesheet loaded. Data crunching beginning.\n')

# Loop through sheets
timesheet_data = {}
for name in timesheet.sheetnames:
    # Ignore the totals sheet
    if(name == TOTAL_SHEET_NAME):
        continue

    sheet = timesheet.get_sheet_by_name(name)

    # If this is January, get the year
    if(name == 'January'):
        cur_year = sheet[YEAR_CELL].value

    print('{} {}'.format(name, cur_year))
    print('==================================')

    # Loop variables
    month_data = {
        'month': name,
        'total_hours': 0,
        'days_worked': []
    }
    cur_po_row = 0

    # Loop through rows
    po = sheet.cell(column=PO_COLUMN, row=PO_START_ROW+cur_po_row).value
    while(po != HOLIDAY and po != ALT_HOLIDAY and po != ALT_ALT_HOLIDAY):
        # Loop variables
        need_to_sort = False

        # Ignore blank rows
        if(po is not None):

            # Check if this row has a '(5)' or '(6)' to indicate out of town
            if(po.find('(5)') is not -1 or po.find('(6)') is not -1):
                # This is an out of town row. Lets move through the columns
                # and count hours and days
                cur_row = PO_START_ROW + cur_po_row

                for cur_col in range(DAYS_START_COLUMN, DAYS_END_COLUMN+1):
                    # Check if this cell is blank, if so, continue
                    hours = sheet.cell(row=cur_row, column=cur_col).value
                    if(hours is None):
                        continue

                    # Ok. Now add the hours to the count.
                    month_data['total_hours'] += int(hours)

                    # Check if we need to add this day to the list
                    cur_day = sheet.cell(row=DAY_NUMBER_ROW, column=cur_col).value
                    if(month_data['days_worked'].count(cur_day) is 0):
                        month_data['days_worked'].append(cur_day)
                        need_to_sort = True
                    
                # Done looping print results
                print('{}: Total Hours:{} Total Days:{}'.format(
                    po, month_data['total_hours'], len(month_data['days_worked'])
                ))

        # Ok. We are done with this row, check if we should sort the days
        if(need_to_sort):
            month_data['days_worked'].sort()

        # Iterate
        cur_po_row += 1
        po = sheet.cell(column=PO_COLUMN, row=PO_START_ROW+cur_po_row).value

    # If we had no data, report so.
    if(month_data['total_hours'] is 0):
        print('No out of town hours this month.')

    # Calculate total days worked
    month_data['total_days'] = len(month_data['days_worked'])

    # Add this months data to the list
    timesheet_data[name] = month_data
    print('')

# Now that we have all of the data, lets add it to the results sheet
print('Timesheet read complete! Beginning results reporting...')
results = openpyxl.load_workbook(results_file)

# Check if we have a sheet for this year. If not, create the sheet and apply
# the formatting.
if(results.sheetnames.count(str(cur_year)) is 0):
    make_new_year_page(results, str(cur_year))

results_sheet = results.get_sheet_by_name(str(cur_year))

# Now add the data to the sheet
month_mapping = {
    'January': 3, 'February': 4, 'March': 5, 'April': 6, 'May': 7, 'June': 8,
    'July': 9, 'August': 10, 'September': 11, 'October': 12, 'November': 13,
    'December': 14
}
for month in timesheet_data:
    row = month_mapping[month]
    cur_month = timesheet_data[month]
    results_sheet.cell(row=row, column=1).value = cur_month['month']
    results_sheet.cell(row=row, column=2).value = cur_month['total_hours']
    results_sheet.cell(row=row, column=3).value = cur_month['total_days']

# Done!
print('Done!')
results.save(results_file)